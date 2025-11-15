#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Portfolio FIFO Analyzer
=======================
Automatische Analyse von Portfolio-CSVs mit FIFO-Prinzip
Erstellt detaillierte Reports und interaktive Visualisierungen

Unterstützte CSV-Formate:
- Semicolon (;) und Comma (,) Trennzeichen
- Verschiedene Datumsformate
- Deutsche und englische Spaltennamen
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

class PortfolioFIFOAnalyzer:
    """
    Automatische FIFO-Analyse für Portfolio-CSVs
    """
    
    def __init__(self, csv_file_path, current_price=None, currency="EUR"):
        self.csv_file_path = csv_file_path
        self.current_price = current_price
        self.currency = currency
        self.company_name = ""
        self.raw_data = None
        self.transactions = None
        self.portfolio = []
        self.analysis_results = {}

        # Dividend tracking
        self.total_dividends = 0
        self.dividend_transactions = []

        # Analyse ausführen
        self._load_and_clean_data()
        self._perform_fifo_analysis()
        self._calculate_summary_stats()

    def _load_and_clean_data(self):
        """
        Lädt und bereinigt die CSV-Daten

        - CSV-Datei einlesen mit automatischer Trennzeichen-Erkennung
        - Spaltennamen normalisieren (Deutsch/Englisch)
        - Zahlenformate konvertieren (Komma → Punkt)
        - Datumsformate parsen
        - Firmenname extrahieren
        """
        # Überprüfe ob Datei existiert
        if not Path(self.csv_file_path).exists():
            raise FileNotFoundError(f"CSV-Datei nicht gefunden: {self.csv_file_path}")

        # Automatische Trennzeichen-Erkennung
        delimiter = self._detect_delimiter()

        # CSV einlesen
        try:
            self.raw_data = pd.read_csv(
                self.csv_file_path,
                delimiter=delimiter,
                encoding='utf-8'
            )
        except UnicodeDecodeError:
            # Fallback auf Latin-1 Encoding
            self.raw_data = pd.read_csv(
                self.csv_file_path,
                delimiter=delimiter,
                encoding='latin-1'
            )

        # Prüfe ob DataFrame leer ist
        if self.raw_data.empty:
            raise ValueError("CSV-Datei ist leer")

        # Spaltennamen normalisieren (Deutsch → Englisch)
        self._normalize_column_names()

        # Überprüfe ob alle erforderlichen Spalten vorhanden sind
        required_columns = ['Date', 'Type', 'Shares', 'Price', 'Amount']
        missing_columns = [col for col in required_columns if col not in self.raw_data.columns]
        if missing_columns:
            raise ValueError(f"Fehlende Spalten: {', '.join(missing_columns)}")

        # Kopie für Transaktionen erstellen
        self.transactions = self.raw_data.copy()

        # Zahlenformate konvertieren (deutsche Kommas → Punkte)
        self._convert_number_formats()

        # Datumsformate parsen
        self._parse_dates()

        # Firmenname extrahieren
        self._extract_company_name()

        # NaN-Werte in numerischen Spalten mit 0 ersetzen
        numeric_columns = ['Shares', 'Price', 'Amount', 'Tax']
        for col in numeric_columns:
            if col in self.transactions.columns:
                self.transactions[col] = self.transactions[col].fillna(0)

    def _detect_delimiter(self):
        """Erkennt automatisch das Trennzeichen der CSV-Datei"""
        with open(self.csv_file_path, 'r', encoding='utf-8') as f:
            first_line = f.readline()

        # Zähle potenzielle Trennzeichen
        delimiters = {
            ';': first_line.count(';'),
            ',': first_line.count(','),
            '\t': first_line.count('\t'),
            '|': first_line.count('|')
        }

        # Wähle das häufigste Trennzeichen
        delimiter = max(delimiters, key=delimiters.get)

        # Fallback auf Komma wenn kein Trennzeichen gefunden
        if delimiters[delimiter] == 0:
            delimiter = ','

        return delimiter

    def _normalize_column_names(self):
        """Normalisiert Spaltennamen (Deutsch → Englisch)"""
        # Mapping von deutschen zu englischen Spaltennamen
        column_mapping = {
            # Datum
            'datum': 'Date',
            'date': 'Date',
            'datetime': 'DateTime',

            # Typ
            'typ': 'Type',
            'type': 'Type',

            # Aktienanzahl
            'anzahl': 'Shares',
            'shares': 'Shares',
            'aktien': 'Shares',

            # Preis
            'preis': 'Price',
            'price': 'Price',
            'kurs': 'Price',

            # Betrag
            'betrag': 'Amount',
            'amount': 'Amount',
            'summe': 'Amount',

            # Steuern
            'steuer': 'Tax',
            'tax': 'Tax',
            'steuern': 'Tax',

            # Gebühren
            'fee': 'Fee',
            'gebühr': 'Fee',
            'gebühren': 'Fee',

            # Realisierte Gewinne
            'realizedgains': 'RealizedGains',
            'gewinn': 'RealizedGains',

            # Firmenname
            'company': 'Company',
            'unternehmen': 'Company',
            'holdingname': 'Company',
            'firma': 'Company',

            # Zeit
            'time': 'Time',
            'zeit': 'Time',
        }

        # Spaltennamen in Kleinbuchstaben umwandeln und mappen
        new_columns = {}
        for col in self.raw_data.columns:
            col_lower = col.lower()
            if col_lower in column_mapping:
                new_columns[col] = column_mapping[col_lower]

        # Umbenennen
        if new_columns:
            self.raw_data.rename(columns=new_columns, inplace=True)

    def _convert_number_formats(self):
        """Konvertiert deutsche Zahlenformate (Komma) zu Punkten"""
        numeric_columns = ['Shares', 'Price', 'Amount', 'Tax', 'Fee', 'RealizedGains']

        for col in numeric_columns:
            if col not in self.transactions.columns:
                continue

            # Überprüfe ob Spalte bereits numerisch ist
            if pd.api.types.is_numeric_dtype(self.transactions[col]):
                continue

            # Konvertiere String-Werte
            def convert_number(value):
                if pd.isna(value):
                    return 0

                # Zu String konvertieren
                str_value = str(value).strip()

                # Leere Strings → 0
                if not str_value:
                    return 0

                # Entferne Tausenderpunkte und ersetze Komma durch Punkt
                str_value = str_value.replace('.', '')  # Tausenderpunkt entfernen
                str_value = str_value.replace(',', '.')  # Komma → Punkt

                try:
                    return float(str_value)
                except ValueError:
                    return 0

            self.transactions[col] = self.transactions[col].apply(convert_number)

    def _parse_dates(self):
        """Parst verschiedene Datumsformate"""
        # Primäre Datumsspalte bestimmen
        date_column = None
        if 'DateTime' in self.transactions.columns:
            date_column = 'DateTime'
        elif 'Date' in self.transactions.columns:
            date_column = 'Date'

        if not date_column:
            return

        # Verschiedene Datumsformate versuchen
        date_formats = [
            '%Y-%m-%d',           # ISO: 2024-01-15
            '%d.%m.%Y',           # Deutsch: 15.01.2024
            '%Y-%m-%dT%H:%M:%S.%fZ',  # ISO DateTime: 2024-01-15T10:30:00.000Z
            '%d/%m/%Y',           # 15/01/2024
            '%m/%d/%Y',           # 01/15/2024
        ]

        # Versuche pandas automatisches Parsing
        try:
            self.transactions[date_column] = pd.to_datetime(
                self.transactions[date_column],
                dayfirst=True  # Für europäische Datumsformate
            )
        except:
            # Fallback: Versuche verschiedene Formate
            for fmt in date_formats:
                try:
                    self.transactions[date_column] = pd.to_datetime(
                        self.transactions[date_column],
                        format=fmt
                    )
                    break
                except:
                    continue

        # Benenne zu 'Date' um für Konsistenz
        if date_column == 'DateTime':
            self.transactions['Date'] = self.transactions['DateTime']

    def _extract_company_name(self):
        """Extrahiert den Firmennamen aus der ersten Zeile"""
        if 'Company' in self.transactions.columns:
            # Erste nicht-leere Firma nehmen
            company_series = self.transactions['Company'].dropna()
            if not company_series.empty:
                self.company_name = str(company_series.iloc[0])

        # Fallback
        if not self.company_name:
            self.company_name = "Unknown Company"

    def _perform_fifo_analysis(self):
        """
        Führt die FIFO-Analyse durch

        - Transaktionen chronologisch sortieren
        - FIFO-Logik anwenden (First-In-First-Out)
        - Realisierte Gewinne/Verluste berechnen
        - Portfolio-Positionen tracken
        """
        # Transaktionen nach Datum sortieren
        self.transactions = self.transactions.sort_values('Date').reset_index(drop=True)

        # Portfolio-Queue für FIFO (Liste von Kauf-Positionen)
        self.portfolio = []

        # Tracking für realisierte Gewinne
        self.realized_gains = []
        self.total_realized_gains = 0
        self.total_taxes = 0

        # Transaktionen durchgehen
        for idx, row in self.transactions.iterrows():
            transaction_type = str(row['Type']).strip().lower()

            # Normalisiere Transaktionstypen (deutsch/englisch)
            if transaction_type in ['buy', 'kauf', 'b']:
                self._process_buy(row)
            elif transaction_type in ['sell', 'verkauf', 'sale', 's']:
                self._process_sell(row)
            elif transaction_type in ['dividend', 'dividende', 'div', 'd']:
                self._process_dividend(row)
            else:
                # Ignoriere unbekannte Transaktionstypen
                continue

    def _process_buy(self, transaction):
        """Verarbeitet eine Kauf-Transaktion"""
        shares = float(transaction['Shares'])
        price = float(transaction['Price'])
        date = transaction['Date']

        # Ignoriere Zero-Share Transaktionen
        if shares == 0:
            return

        # Füge zur Portfolio-Queue hinzu
        self.portfolio.append({
            'date': date,
            'shares': shares,
            'price': price,
            'cost_basis': shares * price
        })

    def _process_sell(self, transaction):
        """Verarbeitet eine Verkaufs-Transaktion (FIFO)"""
        shares_to_sell = float(transaction['Shares'])
        sell_price = float(transaction['Price'])
        sell_amount = float(transaction['Amount'])
        tax = float(transaction.get('Tax', 0))
        date = transaction['Date']

        # Ignoriere Zero-Share Transaktionen
        if shares_to_sell == 0:
            return

        # Überprüfe ob genug Aktien vorhanden sind
        total_shares_available = sum(pos['shares'] for pos in self.portfolio)
        if shares_to_sell > total_shares_available:
            raise ValueError(
                f"Überverkauf am {date}: Versucht {shares_to_sell} Aktien zu verkaufen, "
                f"aber nur {total_shares_available} verfügbar"
            )

        # FIFO: Verkaufe die ältesten Aktien zuerst
        shares_remaining_to_sell = shares_to_sell
        cost_basis_sold = 0
        positions_to_remove = []

        for i, position in enumerate(self.portfolio):
            if shares_remaining_to_sell <= 0:
                break

            if position['shares'] <= shares_remaining_to_sell:
                # Gesamte Position verkaufen
                shares_sold_from_position = position['shares']
                cost_basis_sold += position['cost_basis']
                shares_remaining_to_sell -= position['shares']
                positions_to_remove.append(i)
            else:
                # Teilweise verkaufen
                shares_sold_from_position = shares_remaining_to_sell
                cost_per_share = position['price']
                cost_basis_sold += shares_sold_from_position * cost_per_share

                # Aktualisiere Position
                position['shares'] -= shares_sold_from_position
                position['cost_basis'] = position['shares'] * position['price']
                shares_remaining_to_sell = 0

        # Entferne vollständig verkaufte Positionen (rückwärts um Indizes nicht zu verschieben)
        for i in reversed(positions_to_remove):
            self.portfolio.pop(i)

        # Berechne realisierten Gewinn
        gross_gain = sell_amount - cost_basis_sold
        net_gain = gross_gain - tax

        # Speichere Gewinn-Details
        self.realized_gains.append({
            'date': date,
            'shares': shares_to_sell,
            'sell_price': sell_price,
            'cost_basis': cost_basis_sold,
            'gross_gain': gross_gain,
            'tax': tax,
            'net_gain': net_gain
        })

        # Akkumuliere Gesamt-Werte
        self.total_realized_gains += gross_gain
        self.total_taxes += tax

    def _process_dividend(self, transaction):
        """Verarbeitet eine Dividenden-Transaktion"""
        date = transaction['Date']
        amount = float(transaction['Amount'])
        tax = float(transaction.get('Tax', 0))
        shares = float(transaction.get('Shares', 0))  # Optional: Anzahl Aktien bei Dividende

        # Netto-Dividende (nach Steuern)
        net_dividend = amount - tax

        # Speichere Dividenden-Details
        self.dividend_transactions.append({
            'date': date,
            'amount': amount,
            'tax': tax,
            'net_amount': net_dividend,
            'shares': shares
        })

        # Akkumuliere Gesamt-Dividenden (brutto)
        self.total_dividends += amount

    def _calculate_summary_stats(self):
        """
        Berechnet die zusammenfassenden Statistiken

        - Gesamt-Investition berechnen
        - Gesamt-Entnahmen berechnen
        - Realisierte und unrealisierte Gewinne
        - Rendite in Prozent
        - Netto-Cashflow
        """
        # Gesamt-Investition (alle Käufe)
        buy_transactions = self.transactions[
            self.transactions['Type'].str.lower().isin(['buy', 'kauf', 'b'])
        ]
        total_invested = buy_transactions['Amount'].sum()

        # Gesamt-Entnahmen (alle Verkäufe)
        sell_transactions = self.transactions[
            self.transactions['Type'].str.lower().isin(['sell', 'verkauf', 'sale', 's'])
        ]
        total_withdrawn = sell_transactions['Amount'].sum()

        # Verbleibende Aktien und Kostenbasis
        remaining_shares = sum(pos['shares'] for pos in self.portfolio)
        remaining_cost_basis = sum(pos['cost_basis'] for pos in self.portfolio)

        # Unrealisierte Gewinne (nur wenn current_price gegeben ist)
        unrealized_gains = 0
        current_value = 0
        if self.current_price and remaining_shares > 0:
            current_value = remaining_shares * self.current_price
            unrealized_gains = current_value - remaining_cost_basis

        # Gesamt-Gewinne (realisiert + unrealisiert)
        total_gains = self.total_realized_gains + unrealized_gains

        # Netto-Cashflow (was wurde entnommen - was wurde eingezahlt)
        net_cashflow = total_withdrawn - total_invested

        # Netto realisierte Gewinne (nach Steuern)
        net_realized_gains = self.total_realized_gains - self.total_taxes

        # Dividenden-Statistiken
        dividend_tax = sum(div['tax'] for div in self.dividend_transactions)
        net_dividends = self.total_dividends - dividend_tax

        # Gesamt-Gewinne inkl. Dividenden (realisiert + unrealisiert + dividenden)
        total_gains_incl_div = total_gains + self.total_dividends

        # Gesamt-Rendite in Prozent (mit und ohne Dividenden)
        if total_invested > 0:
            total_return_pct = (total_gains / total_invested) * 100
            total_return_incl_div_pct = (total_gains_incl_div / total_invested) * 100
        else:
            total_return_pct = 0
            total_return_incl_div_pct = 0

        # Speichere alle Statistiken
        self.analysis_results = {
            'total_invested': float(total_invested),
            'total_withdrawn': float(total_withdrawn),
            'total_realized_gains': float(self.total_realized_gains),
            'total_taxes': float(self.total_taxes),
            'net_realized_gains': float(net_realized_gains),
            'unrealized_gains': float(unrealized_gains),
            'total_gains': float(total_gains),
            'net_cashflow': float(net_cashflow),
            'total_return_pct': float(total_return_pct),
            'remaining_shares': float(remaining_shares),
            'remaining_cost_basis': float(remaining_cost_basis),
            'current_value': float(current_value),
            # Dividenden
            'total_dividends': float(self.total_dividends),
            'dividend_tax': float(dividend_tax),
            'net_dividends': float(net_dividends),
            'dividend_count': len(self.dividend_transactions),
            'total_gains_incl_dividends': float(total_gains_incl_div),
            'total_return_incl_dividends_pct': float(total_return_incl_div_pct)
        }

    def print_summary_report(self):
        """Detaillierter Zusammenfassungsbericht"""
        print(f"\n{'='*60}")
        print(f"🚀 {self.company_name.upper()} PORTFOLIO-ANALYSE (FIFO)")
        print(f"{'='*60}")
        
        print(f"\n💰 INVESTITIONS-ÜBERSICHT:")
        print(f"   Gesamt eingezahlt: {self.analysis_results['total_invested']:,.2f} {self.currency}")
        print(f"   Gesamt entnommen:  {self.analysis_results['total_withdrawn']:,.2f} {self.currency}")
        
        print(f"\n📈 REALISIERTE GEWINNE:")
        print(f"   Brutto-Gewinne:    {self.analysis_results['total_realized_gains']:,.2f} {self.currency}")
        print(f"   Steuern gezahlt:   {self.analysis_results['total_taxes']:,.2f} {self.currency}")
        print(f"   Netto-Gewinne:     {self.analysis_results['net_realized_gains']:,.2f} {self.currency}")

        # Dividenden anzeigen, falls vorhanden
        if self.analysis_results['dividend_count'] > 0:
            print(f"\n💵 DIVIDENDEN:")
            print(f"   Anzahl:            {self.analysis_results['dividend_count']}")
            print(f"   Brutto-Dividenden: {self.analysis_results['total_dividends']:,.2f} {self.currency}")
            print(f"   Steuern:           {self.analysis_results['dividend_tax']:,.2f} {self.currency}")
            print(f"   Netto-Dividenden:  {self.analysis_results['net_dividends']:,.2f} {self.currency}")

        print(f"\n🎯 GESAMTERGEBNIS:")
        print(f"   Gesamtgewinn:      {self.analysis_results['total_gains']:,.2f} {self.currency}")
        if self.analysis_results['dividend_count'] > 0:
            print(f"   Inkl. Dividenden:  {self.analysis_results['total_gains_incl_dividends']:,.2f} {self.currency}")
        print(f"   Netto-Cashflow:    {self.analysis_results['net_cashflow']:,.2f} {self.currency}")
        print(f"   Gesamtrendite:     {self.analysis_results['total_return_pct']:,.1f}%")
        if self.analysis_results['dividend_count'] > 0:
            print(f"   Inkl. Dividenden:  {self.analysis_results['total_return_incl_dividends_pct']:,.1f}%")

    def generate_html_report(self, output_file="output/portfolio_report.html"):
        """
        Generiert einen interaktiven HTML-Report

        Args:
            output_file (str): Pfad für die Output-HTML-Datei
        """
        from datetime import datetime

        # Erstelle output-Verzeichnis falls nicht vorhanden
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        # Formatierte Werte für HTML
        stats = self.analysis_results
        company = self.company_name

        # Datum für Report
        report_date = datetime.now().strftime("%d.%m.%Y %H:%M")

        # Bereite Transaktionsdaten für die Tabelle vor
        transactions_html = ""
        for idx, row in self.transactions.iterrows():
            date_str = row['Date'].strftime("%d.%m.%Y") if hasattr(row['Date'], 'strftime') else str(row['Date'])
            trans_type = row['Type']
            shares = row['Shares']
            price = row['Price']
            amount = row['Amount']
            tax = row.get('Tax', 0)

            # Transaktionstyp und Emoji bestimmen
            trans_type_lower = str(trans_type).lower()
            if trans_type_lower in ['buy', 'kauf', 'b']:
                type_class = "buy"
                type_emoji = "🟢"
            elif trans_type_lower in ['dividend', 'dividende', 'div', 'd']:
                type_class = "dividend"
                type_emoji = "💵"
            else:
                type_class = "sell"
                type_emoji = "🔴"

            transactions_html += f"""
                <tr class="{type_class}">
                    <td>{date_str}</td>
                    <td>{type_emoji} {trans_type}</td>
                    <td>{shares:,.0f}</td>
                    <td>{price:,.2f} {self.currency}</td>
                    <td>{amount:,.2f} {self.currency}</td>
                    <td>{tax:,.2f} {self.currency}</td>
                </tr>
            """

        # Bereite Portfolio-Positionen vor
        portfolio_html = ""
        for pos in self.portfolio:
            date_str = pos['date'].strftime("%d.%m.%Y") if hasattr(pos['date'], 'strftime') else str(pos['date'])
            shares = pos['shares']
            price = pos['price']
            cost_basis = pos['cost_basis']

            if self.current_price:
                current_value = shares * self.current_price
                unrealized = current_value - cost_basis
                unrealized_pct = (unrealized / cost_basis * 100) if cost_basis > 0 else 0
                unrealized_class = "positive" if unrealized >= 0 else "negative"

                portfolio_html += f"""
                    <tr>
                        <td>{date_str}</td>
                        <td>{shares:,.2f}</td>
                        <td>{price:,.2f} {self.currency}</td>
                        <td>{cost_basis:,.2f} {self.currency}</td>
                        <td>{current_value:,.2f} {self.currency}</td>
                        <td class="{unrealized_class}">{unrealized:+,.2f} {self.currency} ({unrealized_pct:+.1f}%)</td>
                    </tr>
                """
            else:
                portfolio_html += f"""
                    <tr>
                        <td>{date_str}</td>
                        <td>{shares:,.2f}</td>
                        <td>{price:,.2f} {self.currency}</td>
                        <td>{cost_basis:,.2f} {self.currency}</td>
                        <td colspan="2">-</td>
                    </tr>
                """

        # HTML Template
        html_content = f"""<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{company} - Portfolio FIFO Analyse</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            color: #333;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
        }}
        .header {{
            background: white;
            border-radius: 15px;
            padding: 30px;
            margin-bottom: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        h1 {{
            color: #667eea;
            margin-bottom: 10px;
        }}
        .subtitle {{
            color: #666;
            font-size: 14px;
        }}
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .metric-card {{
            background: white;
            border-radius: 15px;
            padding: 25px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            transition: transform 0.3s;
        }}
        .metric-card:hover {{
            transform: translateY(-5px);
        }}
        .metric-label {{
            color: #666;
            font-size: 14px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 10px;
        }}
        .metric-value {{
            font-size: 32px;
            font-weight: bold;
            color: #333;
            margin-bottom: 5px;
        }}
        .metric-value.positive {{
            color: #10b981;
        }}
        .metric-value.negative {{
            color: #ef4444;
        }}
        .metric-subtext {{
            font-size: 12px;
            color: #999;
        }}
        .section {{
            background: white;
            border-radius: 15px;
            padding: 30px;
            margin-bottom: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        h2 {{
            color: #667eea;
            margin-bottom: 20px;
            font-size: 24px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #f0f0f0;
        }}
        th {{
            background: #f8f9fa;
            font-weight: 600;
            color: #667eea;
        }}
        tr.buy {{
            background: #f0fdf4;
        }}
        tr.sell {{
            background: #fef2f2;
        }}
        tr.dividend {{
            background: #eff6ff;
            font-style: italic;
        }}
        .positive {{
            color: #10b981;
            font-weight: bold;
        }}
        .negative {{
            color: #ef4444;
            font-weight: bold;
        }}
        .chart-container {{
            position: relative;
            height: 400px;
            margin: 20px 0;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 {company} - Portfolio FIFO Analyse</h1>
            <p class="subtitle">Erstellt am: {report_date}</p>
        </div>

        <div class="metrics-grid">
            <div class="metric-card">
                <div class="metric-label">💰 Gesamt Investiert</div>
                <div class="metric-value">{stats['total_invested']:,.2f} {self.currency}</div>
                <div class="metric-subtext">Alle Käufe</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">💸 Gesamt Entnommen</div>
                <div class="metric-value">{stats['total_withdrawn']:,.2f} {self.currency}</div>
                <div class="metric-subtext">Alle Verkäufe</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">📈 Realisierte Gewinne</div>
                <div class="metric-value {'positive' if stats['total_realized_gains'] >= 0 else 'negative'}">
                    {stats['total_realized_gains']:+,.2f} {self.currency}
                </div>
                <div class="metric-subtext">Brutto</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">🏦 Gezahlte Steuern</div>
                <div class="metric-value">{stats['total_taxes']:,.2f} {self.currency}</div>
                <div class="metric-subtext">Abgeltungssteuer</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">💵 Netto-Gewinne</div>
                <div class="metric-value {'positive' if stats['net_realized_gains'] >= 0 else 'negative'}">
                    {stats['net_realized_gains']:+,.2f} {self.currency}
                </div>
                <div class="metric-subtext">Nach Steuern</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">💎 Unrealisierte Gewinne</div>
                <div class="metric-value {'positive' if stats['unrealized_gains'] >= 0 else 'negative'}">
                    {stats['unrealized_gains']:+,.2f} {self.currency}
                </div>
                <div class="metric-subtext">Aktuelle Positionen</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">🎯 Gesamt-Gewinne</div>
                <div class="metric-value {'positive' if stats['total_gains'] >= 0 else 'negative'}">
                    {stats['total_gains']:+,.2f} {self.currency}
                </div>
                <div class="metric-subtext">Realisiert + Unrealisiert</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">💰 Netto-Cashflow</div>
                <div class="metric-value {'positive' if stats['net_cashflow'] >= 0 else 'negative'}">
                    {stats['net_cashflow']:+,.2f} {self.currency}
                </div>
                <div class="metric-subtext">Entnommen - Investiert</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">📊 Gesamt-Rendite</div>
                <div class="metric-value {'positive' if stats['total_return_pct'] >= 0 else 'negative'}">
                    {stats['total_return_pct']:+,.1f}%
                </div>
                <div class="metric-subtext">ROI</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">🔢 Verbleibende Aktien</div>
                <div class="metric-value">{stats['remaining_shares']:,.0f}</div>
                <div class="metric-subtext">Im Portfolio</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">📉 Kostenbasis</div>
                <div class="metric-value">{stats['remaining_cost_basis']:,.2f} {self.currency}</div>
                <div class="metric-subtext">Verbleibende Positionen</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">💰 Aktueller Wert</div>
                <div class="metric-value">{stats['current_value']:,.2f} {self.currency}</div>
                <div class="metric-subtext">Bei aktuellem Kurs</div>
            </div>"""

        # Dividenden-Metriken hinzufügen, falls vorhanden
        if stats['dividend_count'] > 0:
            html_content += f"""
            <div class="metric-card" style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);">
                <div class="metric-label">💵 Dividenden (Brutto)</div>
                <div class="metric-value" style="color: white;">{stats['total_dividends']:,.2f} {self.currency}</div>
                <div class="metric-subtext" style="color: rgba(255,255,255,0.8);">{stats['dividend_count']} Zahlungen</div>
            </div>

            <div class="metric-card" style="background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);">
                <div class="metric-label">💵 Dividenden (Netto)</div>
                <div class="metric-value" style="color: white;">{stats['net_dividends']:,.2f} {self.currency}</div>
                <div class="metric-subtext" style="color: rgba(255,255,255,0.8);">Nach Steuern</div>
            </div>

            <div class="metric-card" style="background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);">
                <div class="metric-label">🎯 Rendite (inkl. Div.)</div>
                <div class="metric-value {'positive' if stats['total_return_incl_dividends_pct'] >= 0 else 'negative'}" style="color: white;">
                    {stats['total_return_incl_dividends_pct']:+,.1f}%
                </div>
                <div class="metric-subtext" style="color: rgba(255,255,255,0.8);">Mit Dividenden</div>
            </div>"""

        html_content += """
        </div>

        <div class="section">
            <h2>📋 Transaktionshistorie</h2>
            <table>
                <thead>
                    <tr>
                        <th>Datum</th>
                        <th>Typ</th>
                        <th>Aktien</th>
                        <th>Preis</th>
                        <th>Betrag</th>
                        <th>Steuern</th>
                    </tr>
                </thead>
                <tbody>
                    {transactions_html}
                </tbody>
            </table>
        </div>

        <div class="section">
            <h2>💼 Aktuelle Portfolio-Positionen</h2>
            <table>
                <thead>
                    <tr>
                        <th>Kaufdatum</th>
                        <th>Aktien</th>
                        <th>Kaufpreis</th>
                        <th>Kostenbasis</th>
                        <th>Aktueller Wert</th>
                        <th>Unrealisierter Gewinn</th>
                    </tr>
                </thead>
                <tbody>
                    {portfolio_html}
                </tbody>
            </table>
        </div>

        <div class="section">
            <p style="text-align: center; color: #999; font-size: 12px;">
                ⚠️ <strong>Rechtlicher Hinweis:</strong> Dieses Tool dient nur zur Unterstützung bei der Portfolioanalyse.
                Es ersetzt keine professionelle Steuerberatung. Alle Berechnungen sollten vor steuerlichen Entscheidungen
                von einem Steuerberater überprüft werden.
            </p>
        </div>
    </div>
</body>
</html>"""

        # Schreibe HTML-Datei
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return output_file

    def generate_excel_report(self, output_file="output/portfolio_report.xlsx"):
        """
        Generiert einen detaillierten Excel-Report

        Args:
            output_file (str): Pfad für die Output-Excel-Datei
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("openpyxl ist nicht installiert. Bitte installieren: pip install openpyxl")

        from datetime import datetime

        # Erstelle output-Verzeichnis falls nicht vorhanden
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        # Erstelle Workbook
        wb = Workbook()

        # === Sheet 1: Zusammenfassung ===
        ws_summary = wb.active
        ws_summary.title = "Zusammenfassung"

        # Header-Stil
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Titel
        ws_summary['A1'] = f"{self.company_name} - Portfolio FIFO Analyse"
        ws_summary['A1'].font = Font(bold=True, size=16, color="667EEA")
        ws_summary['A2'] = f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws_summary['A2'].font = Font(size=10, color="666666")

        # Statistiken
        stats = self.analysis_results
        row = 4

        # Investitions-Übersicht
        ws_summary[f'A{row}'] = "INVESTITIONS-ÜBERSICHT"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        data_rows = [
            ("Gesamt eingezahlt", stats['total_invested']),
            ("Gesamt entnommen", stats['total_withdrawn']),
            ("", ""),
            ("REALISIERTE GEWINNE", ""),
            ("Brutto-Gewinne", stats['total_realized_gains']),
            ("Steuern gezahlt", stats['total_taxes']),
            ("Netto-Gewinne", stats['net_realized_gains']),
            ("", ""),
            ("UNREALISIERTE GEWINNE", ""),
            ("Aktueller Wert", stats['current_value']),
            ("Kostenbasis", stats['remaining_cost_basis']),
            ("Unrealisierte Gewinne", stats['unrealized_gains']),
        ]

        # Dividenden hinzufügen, falls vorhanden
        if stats['dividend_count'] > 0:
            data_rows.extend([
                ("", ""),
                ("DIVIDENDEN", ""),
                ("Anzahl Zahlungen", stats['dividend_count']),
                ("Brutto-Dividenden", stats['total_dividends']),
                ("Steuern", stats['dividend_tax']),
                ("Netto-Dividenden", stats['net_dividends']),
            ])

        data_rows.extend([
            ("", ""),
            ("GESAMTERGEBNIS", ""),
            ("Gesamt-Gewinne", stats['total_gains']),
        ])

        if stats['dividend_count'] > 0:
            data_rows.append(("Inkl. Dividenden", stats['total_gains_incl_dividends']))

        data_rows.extend([
            ("Netto-Cashflow", stats['net_cashflow']),
            ("Gesamt-Rendite (%)", stats['total_return_pct']),
        ])

        if stats['dividend_count'] > 0:
            data_rows.append(("Rendite inkl. Div. (%)", stats['total_return_incl_dividends_pct']))

        for label, value in data_rows:
            if label in ["INVESTITIONS-ÜBERSICHT", "REALISIERTE GEWINNE", "UNREALISIERTE GEWINNE", "DIVIDENDEN", "GESAMTERGEBNIS"]:
                ws_summary[f'A{row}'] = label
                ws_summary[f'A{row}'].font = Font(bold=True)
            elif label:
                ws_summary[f'A{row}'] = label
                if isinstance(value, (int, float)):
                    ws_summary[f'B{row}'] = value
                    if '%' in label:
                        ws_summary[f'B{row}'].number_format = '0.00"%"'
                    else:
                        ws_summary[f'B{row}'].number_format = '#,##0.00 "' + self.currency + '"'

                    # Farbcodierung
                    if value > 0 and 'Gewinn' in label:
                        ws_summary[f'B{row}'].font = Font(color="10B981", bold=True)
                    elif value < 0:
                        ws_summary[f'B{row}'].font = Font(color="EF4444", bold=True)
            row += 1

        # Spaltenbreiten anpassen
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        # === Sheet 2: Transaktionen ===
        ws_trans = wb.create_sheet("Transaktionen")

        # Header
        headers = ["Datum", "Typ", "Aktien", "Preis", "Betrag", "Steuern"]
        for col, header in enumerate(headers, 1):
            cell = ws_trans.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Daten
        for row_idx, (idx, row) in enumerate(self.transactions.iterrows(), 2):
            date_val = row['Date'].strftime("%d.%m.%Y") if hasattr(row['Date'], 'strftime') else str(row['Date'])
            ws_trans.cell(row_idx, 1, date_val)
            ws_trans.cell(row_idx, 2, row['Type'])
            ws_trans.cell(row_idx, 3, row['Shares'])
            ws_trans.cell(row_idx, 4, row['Price']).number_format = '#,##0.00 "' + self.currency + '"'
            ws_trans.cell(row_idx, 5, row['Amount']).number_format = '#,##0.00 "' + self.currency + '"'
            ws_trans.cell(row_idx, 6, row.get('Tax', 0)).number_format = '#,##0.00 "' + self.currency + '"'

            # Farbcodierung
            if str(row['Type']).lower() in ['buy', 'kauf']:
                for col in range(1, 7):
                    ws_trans.cell(row_idx, col).fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
            else:
                for col in range(1, 7):
                    ws_trans.cell(row_idx, col).fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

        # Spaltenbreiten
        for col in range(1, 7):
            ws_trans.column_dimensions[chr(64 + col)].width = 15

        # === Sheet 3: Portfolio-Positionen ===
        ws_portfolio = wb.create_sheet("Portfolio-Positionen")

        # Header
        headers = ["Kaufdatum", "Aktien", "Kaufpreis", "Kostenbasis", "Aktueller Wert", "Unrealisierter Gewinn", "Rendite %"]
        for col, header in enumerate(headers, 1):
            cell = ws_portfolio.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Daten
        for row_idx, pos in enumerate(self.portfolio, 2):
            date_val = pos['date'].strftime("%d.%m.%Y") if hasattr(pos['date'], 'strftime') else str(pos['date'])
            ws_portfolio.cell(row_idx, 1, date_val)
            ws_portfolio.cell(row_idx, 2, pos['shares']).number_format = '#,##0.00'
            ws_portfolio.cell(row_idx, 3, pos['price']).number_format = '#,##0.00 "' + self.currency + '"'
            ws_portfolio.cell(row_idx, 4, pos['cost_basis']).number_format = '#,##0.00 "' + self.currency + '"'

            if self.current_price:
                current_value = pos['shares'] * self.current_price
                unrealized = current_value - pos['cost_basis']
                unrealized_pct = (unrealized / pos['cost_basis'] * 100) if pos['cost_basis'] > 0 else 0

                ws_portfolio.cell(row_idx, 5, current_value).number_format = '#,##0.00 "' + self.currency + '"'
                ws_portfolio.cell(row_idx, 6, unrealized).number_format = '#,##0.00 "' + self.currency + '"'
                ws_portfolio.cell(row_idx, 7, unrealized_pct).number_format = '0.00"%"'

                # Farbcodierung
                if unrealized >= 0:
                    ws_portfolio.cell(row_idx, 6).font = Font(color="10B981", bold=True)
                    ws_portfolio.cell(row_idx, 7).font = Font(color="10B981", bold=True)
                else:
                    ws_portfolio.cell(row_idx, 6).font = Font(color="EF4444", bold=True)
                    ws_portfolio.cell(row_idx, 7).font = Font(color="EF4444", bold=True)

        # Spaltenbreiten
        for col in range(1, 8):
            ws_portfolio.column_dimensions[chr(64 + col)].width = 18

        # Speichere Workbook
        wb.save(output_file)

        return output_file

    def generate_pdf_report(self, output_file="output/portfolio_report.pdf"):
        """
        Generiert einen professionellen PDF-Report

        Args:
            output_file: Pfad für PDF-Ausgabedatei

        Returns:
            str: Pfad zur erstellten PDF-Datei
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.platypus import Image as RLImage
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from datetime import datetime
        import io

        # Sicherstellen, dass das output-Verzeichnis existiert
        import os
        os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)

        # PDF-Dokument erstellen
        doc = SimpleDocTemplate(
            output_file,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12,
            spaceBefore=12
        )

        # Inhalte sammeln
        story = []

        # Titel
        company_name = self.company_name if self.company_name else "Portfolio"
        title = Paragraph(f"📊 {company_name} - FIFO Analyse", title_style)
        story.append(title)

        # Datum
        date_text = Paragraph(
            f"<para align=center>Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}</para>",
            styles['Normal']
        )
        story.append(date_text)
        story.append(Spacer(1, 0.5*cm))

        # Zusammenfassung
        story.append(Paragraph("📈 Zusammenfassung", heading_style))

        stats = self.analysis_results

        summary_data = [
            ['Kennzahl', 'Wert'],
            ['Gesamt eingezahlt', f"{stats['total_invested']:,.2f} {self.currency}"],
            ['Gesamt entnommen', f"{stats['total_withdrawn']:,.2f} {self.currency}"],
            ['Realisierte Gewinne (brutto)', f"{stats['total_realized_gains']:,.2f} {self.currency}"],
            ['Gezahlte Steuern', f"{stats['total_taxes']:,.2f} {self.currency}"],
            ['Realisierte Gewinne (netto)', f"{stats['net_realized_gains']:,.2f} {self.currency}"],
        ]

        # Dividenden hinzufügen, falls vorhanden
        if stats['dividend_count'] > 0:
            summary_data.extend([
                ['Dividenden (brutto)', f"{stats['total_dividends']:,.2f} {self.currency}"],
                ['Dividenden Steuern', f"{stats['dividend_tax']:,.2f} {self.currency}"],
                ['Dividenden (netto)', f"{stats['net_dividends']:,.2f} {self.currency}"],
            ])

        summary_data.extend([
            ['Verbleibende Aktien', f"{stats['remaining_shares']:.0f}"],
            ['Aktueller Wert', f"{stats['current_value']:,.2f} {self.currency}"],
            ['Unrealisierte Gewinne', f"{stats['unrealized_gains']:,.2f} {self.currency}"],
            ['Gesamtgewinn', f"{stats['total_gains']:,.2f} {self.currency}"],
        ])

        if stats['dividend_count'] > 0:
            summary_data.append(['Inkl. Dividenden', f"{stats['total_gains_incl_dividends']:,.2f} {self.currency}"])

        summary_data.extend([
            ['Netto-Cashflow', f"{stats['net_cashflow']:,.2f} {self.currency}"],
            ['Gesamtrendite', f"{stats['total_return_pct']:.1f}%"],
        ])

        if stats['dividend_count'] > 0:
            summary_data.append(['Rendite inkl. Div.', f"{stats['total_return_incl_dividends_pct']:.1f}%"])

        summary_table = Table(summary_data, colWidths=[10*cm, 6*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),  # Highlight last 3 rows
            ('BACKGROUND', (0, -3), (-1, -1), colors.HexColor('#e8f4f8')),
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 1*cm))

        # Transaktionen
        story.append(Paragraph("📋 Transaktionen", heading_style))

        # Transaktionsdaten vorbereiten
        trans_data = [['Datum', 'Typ', 'Aktien', 'Preis', 'Betrag', 'Steuer']]

        for _, row in self.transactions.iterrows():
            trans_data.append([
                row['Date'].strftime('%d.%m.%Y'),
                row['Type'],
                f"{row['Shares']:.0f}",
                f"{row['Price']:.2f}",
                f"{row['Amount']:.2f}",
                f"{row['Tax']:.2f}" if row['Tax'] > 0 else '-'
            ])

        trans_table = Table(trans_data, colWidths=[2.5*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm, 2*cm])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))

        story.append(trans_table)
        story.append(Spacer(1, 1*cm))

        # Portfolio-Positionen
        if self.portfolio and len(self.portfolio) > 0 and self.current_price:
            story.append(Paragraph("💼 Aktuelle Portfolio-Positionen", heading_style))

            pos_data = [['Kaufdatum', 'Kaufpreis', 'Aktien', 'Kosten', 'Aktuell', 'Gewinn', 'Rendite']]

            for pos in self.portfolio:
                if pos['shares'] > 0:
                    current_val = pos['shares'] * self.current_price
                    unrealized = current_val - pos['cost_basis']
                    unrealized_pct = (unrealized / pos['cost_basis'] * 100) if pos['cost_basis'] > 0 else 0

                    pos_data.append([
                        pos['date'].strftime('%d.%m.%Y') if hasattr(pos['date'], 'strftime') else str(pos['date']),
                        f"{pos['price']:.2f}",
                        f"{pos['shares']:.0f}",
                        f"{pos['cost_basis']:.2f}",
                        f"{current_val:.2f}",
                        f"{unrealized:+.2f}",
                        f"{unrealized_pct:+.1f}%"
                    ])

            pos_table = Table(pos_data, colWidths=[2.3*cm, 2*cm, 1.8*cm, 2*cm, 2*cm, 2*cm, 2*cm])
            pos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27ae60')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))

            story.append(pos_table)

        # Fußzeile
        story.append(Spacer(1, 2*cm))
        footer = Paragraph(
            "<para align=center><i>Erstellt mit Portfolio FIFO Analyzer</i></para>",
            styles['Normal']
        )
        story.append(footer)

        # PDF erstellen
        doc.build(story)

        return output_file


def analyze_portfolio_from_csv(csv_file_path, current_price=None, currency="EUR"):
    """
    Hauptfunktion zur Portfolio-Analyse

    Args:
        csv_file_path (str): Pfad zur CSV-Datei
        current_price (float, optional): Aktueller Aktienkurs
        currency (str): Währung (Standard: EUR)

    Returns:
        PortfolioFIFOAnalyzer: Analyzer-Objekt mit allen Ergebnissen
    """
    analyzer = PortfolioFIFOAnalyzer(csv_file_path, current_price, currency)
    analyzer.print_summary_report()
    return analyzer


class MultiAssetPortfolioAnalyzer:
    """
    FIFO-Analyse für Portfolios mit mehreren Aktien

    Verwaltet mehrere PortfolioFIFOAnalyzer Instanzen und erstellt
    aggregierte Reports über alle Assets hinweg.
    """

    def __init__(self, csv_file_path, current_prices=None, currency="EUR"):
        """
        Initialisiert Multi-Asset Analyzer

        Args:
            csv_file_path (str): Pfad zur CSV-Datei (mit Company-Spalte)
                                oder Dictionary {company: csv_path}
            current_prices (dict, optional): {company: price} Dictionary
            currency (str): Währung (Standard: EUR)
        """
        self.currency = currency
        self.current_prices = current_prices or {}
        self.analyzers = {}
        self.companies = []

        # Lade und verarbeite Daten
        if isinstance(csv_file_path, str):
            # Einzelne CSV-Datei mit mehreren Companies
            self._load_from_single_csv(csv_file_path)
        elif isinstance(csv_file_path, dict):
            # Dictionary von CSV-Dateien pro Company
            self._load_from_multiple_csvs(csv_file_path)
        else:
            raise ValueError("csv_file_path muss ein String oder Dictionary sein")

        # Berechne aggregierte Statistiken
        self._calculate_aggregate_stats()

    def _load_from_single_csv(self, csv_file_path):
        """Lädt Daten aus einer CSV-Datei mit mehreren Companies"""
        # CSV einlesen
        delimiter = self._detect_delimiter(csv_file_path)

        try:
            df = pd.read_csv(csv_file_path, delimiter=delimiter, encoding='utf-8')
        except UnicodeDecodeError:
            df = pd.read_csv(csv_file_path, delimiter=delimiter, encoding='latin-1')

        # Spaltennamen normalisieren
        column_mapping = {
            'company': 'Company',
            'unternehmen': 'Company',
            'holdingname': 'Company',
            'firma': 'Company'
        }

        for col in df.columns:
            col_lower = col.lower()
            if col_lower in column_mapping:
                df.rename(columns={col: column_mapping[col_lower]}, inplace=True)

        # Überprüfe ob Company-Spalte existiert
        if 'Company' not in df.columns:
            raise ValueError("CSV-Datei muss eine 'Company' Spalte enthalten für Multi-Asset Analyse")

        # Gruppiere nach Company
        for company in df['Company'].unique():
            if pd.isna(company):
                continue

            company_str = str(company)
            self.companies.append(company_str)

            # Erstelle temporäre CSV für diese Company
            company_df = df[df['Company'] == company].copy()
            temp_csv_path = f"/tmp/portfolio_{company_str.replace(' ', '_')}.csv"
            company_df.to_csv(temp_csv_path, index=False)

            # Erstelle Analyzer für diese Company
            current_price = self.current_prices.get(company_str)
            self.analyzers[company_str] = PortfolioFIFOAnalyzer(
                temp_csv_path,
                current_price,
                self.currency
            )

    def _load_from_multiple_csvs(self, csv_files):
        """Lädt Daten aus mehreren CSV-Dateien"""
        for company, csv_path in csv_files.items():
            self.companies.append(company)
            current_price = self.current_prices.get(company)
            self.analyzers[company] = PortfolioFIFOAnalyzer(
                csv_path,
                current_price,
                self.currency
            )

    def _detect_delimiter(self, csv_file_path):
        """Erkennt automatisch das Trennzeichen der CSV-Datei"""
        with open(csv_file_path, 'r', encoding='utf-8') as f:
            first_line = f.readline()

        delimiters = {
            ';': first_line.count(';'),
            ',': first_line.count(','),
            '\t': first_line.count('\t'),
            '|': first_line.count('|')
        }

        delimiter = max(delimiters, key=delimiters.get)
        if delimiters[delimiter] == 0:
            delimiter = ','

        return delimiter

    def _calculate_aggregate_stats(self):
        """Berechnet aggregierte Statistiken über alle Assets"""
        self.aggregate_stats = {
            'total_invested': 0,
            'total_withdrawn': 0,
            'total_realized_gains': 0,
            'total_taxes': 0,
            'net_realized_gains': 0,
            'unrealized_gains': 0,
            'total_gains': 0,
            'net_cashflow': 0,
            'remaining_shares_value': 0,
            'remaining_cost_basis': 0,
            'num_assets': len(self.companies)
        }

        for analyzer in self.analyzers.values():
            stats = analyzer.analysis_results
            self.aggregate_stats['total_invested'] += stats['total_invested']
            self.aggregate_stats['total_withdrawn'] += stats['total_withdrawn']
            self.aggregate_stats['total_realized_gains'] += stats['total_realized_gains']
            self.aggregate_stats['total_taxes'] += stats['total_taxes']
            self.aggregate_stats['net_realized_gains'] += stats['net_realized_gains']
            self.aggregate_stats['unrealized_gains'] += stats['unrealized_gains']
            self.aggregate_stats['total_gains'] += stats['total_gains']
            self.aggregate_stats['net_cashflow'] += stats['net_cashflow']
            self.aggregate_stats['remaining_shares_value'] += stats['current_value']
            self.aggregate_stats['remaining_cost_basis'] += stats['remaining_cost_basis']

        # Berechne Gesamt-Rendite
        if self.aggregate_stats['total_invested'] > 0:
            self.aggregate_stats['total_return_pct'] = (
                self.aggregate_stats['total_gains'] /
                self.aggregate_stats['total_invested'] * 100
            )
        else:
            self.aggregate_stats['total_return_pct'] = 0

    def print_summary_report(self):
        """Druckt zusammenfassenden Report für alle Assets"""
        print(f"\n{'='*60}")
        print(f"🌐 MULTI-ASSET PORTFOLIO-ANALYSE (FIFO)")
        print(f"{'='*60}")
        print(f"📊 Anzahl Aktien: {self.aggregate_stats['num_assets']}")
        print(f"💰 Währung: {self.currency}")

        print(f"\n💰 GESAMT-INVESTITION:")
        print(f"   Gesamt eingezahlt: {self.aggregate_stats['total_invested']:,.2f} {self.currency}")
        print(f"   Gesamt entnommen:  {self.aggregate_stats['total_withdrawn']:,.2f} {self.currency}")

        print(f"\n📈 REALISIERTE GEWINNE:")
        print(f"   Brutto-Gewinne:    {self.aggregate_stats['total_realized_gains']:,.2f} {self.currency}")
        print(f"   Steuern gezahlt:   {self.aggregate_stats['total_taxes']:,.2f} {self.currency}")
        print(f"   Netto-Gewinne:     {self.aggregate_stats['net_realized_gains']:,.2f} {self.currency}")

        print(f"\n💎 UNREALISIERTE GEWINNE:")
        print(f"   Aktuelle Positionen: {self.aggregate_stats['unrealized_gains']:,.2f} {self.currency}")
        print(f"   Aktueller Wert:      {self.aggregate_stats['remaining_shares_value']:,.2f} {self.currency}")
        print(f"   Kostenbasis:         {self.aggregate_stats['remaining_cost_basis']:,.2f} {self.currency}")

        print(f"\n🎯 GESAMTERGEBNIS:")
        print(f"   Gesamtgewinn:      {self.aggregate_stats['total_gains']:,.2f} {self.currency}")
        print(f"   Netto-Cashflow:    {self.aggregate_stats['net_cashflow']:,.2f} {self.currency}")
        print(f"   Gesamtrendite:     {self.aggregate_stats['total_return_pct']:,.1f}%")

        # Detaillierte Übersicht pro Asset
        print(f"\n{'='*60}")
        print(f"📊 EINZELNE ASSETS:")
        print(f"{'='*60}")

        for company in sorted(self.companies):
            analyzer = self.analyzers[company]
            stats = analyzer.analysis_results

            print(f"\n🔹 {company}")
            print(f"   Investiert:      {stats['total_invested']:>12,.2f} {self.currency}")
            print(f"   Realisiert:      {stats['total_realized_gains']:>12,.2f} {self.currency}")
            print(f"   Unrealisiert:    {stats['unrealized_gains']:>12,.2f} {self.currency}")
            print(f"   Rendite:         {stats['total_return_pct']:>12,.1f}%")

    def generate_html_report(self, output_file="output/multi_asset_report.html"):
        """Generiert HTML-Report für Multi-Asset Portfolio"""
        from datetime import datetime

        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        report_date = datetime.now().strftime("%d.%m.%Y %H:%M")
        stats = self.aggregate_stats

        # Asset-Tabelle erstellen
        assets_html = ""
        for company in sorted(self.companies):
            analyzer = self.analyzers[company]
            astats = analyzer.analysis_results

            return_class = "positive" if astats['total_return_pct'] >= 0 else "negative"

            assets_html += f"""
                <tr>
                    <td><strong>{company}</strong></td>
                    <td>{astats['total_invested']:,.2f} {self.currency}</td>
                    <td>{astats['total_withdrawn']:,.2f} {self.currency}</td>
                    <td class="{'positive' if astats['total_realized_gains'] >= 0 else 'negative'}">
                        {astats['total_realized_gains']:+,.2f} {self.currency}
                    </td>
                    <td class="{'positive' if astats['unrealized_gains'] >= 0 else 'negative'}">
                        {astats['unrealized_gains']:+,.2f} {self.currency}
                    </td>
                    <td class="{'positive' if astats['total_gains'] >= 0 else 'negative'}">
                        {astats['total_gains']:+,.2f} {self.currency}
                    </td>
                    <td class="{return_class}">
                        {astats['total_return_pct']:+,.1f}%
                    </td>
                </tr>
            """

        # HTML Template
        html_content = f"""<!DOCTYPE html>
<html lang="de">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Multi-Asset Portfolio - FIFO Analyse</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            color: #333;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; }}
        .header {{
            background: white;
            border-radius: 15px;
            padding: 30px;
            margin-bottom: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        h1 {{ color: #667eea; margin-bottom: 10px; }}
        .subtitle {{ color: #666; font-size: 14px; }}
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .metric-card {{
            background: white;
            border-radius: 15px;
            padding: 25px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            transition: transform 0.3s;
        }}
        .metric-card:hover {{ transform: translateY(-5px); }}
        .metric-label {{
            color: #666;
            font-size: 14px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 10px;
        }}
        .metric-value {{
            font-size: 32px;
            font-weight: bold;
            color: #333;
            margin-bottom: 5px;
        }}
        .metric-value.positive {{ color: #10b981; }}
        .metric-value.negative {{ color: #ef4444; }}
        .metric-subtext {{ font-size: 12px; color: #999; }}
        .section {{
            background: white;
            border-radius: 15px;
            padding: 30px;
            margin-bottom: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }}
        h2 {{ color: #667eea; margin-bottom: 20px; font-size: 24px; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #f0f0f0; }}
        th {{ background: #f8f9fa; font-weight: 600; color: #667eea; }}
        .positive {{ color: #10b981; font-weight: bold; }}
        .negative {{ color: #ef4444; font-weight: bold; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🌐 Multi-Asset Portfolio - FIFO Analyse</h1>
            <p class="subtitle">Erstellt am: {report_date} | {stats['num_assets']} Assets</p>
        </div>

        <div class="metrics-grid">
            <div class="metric-card">
                <div class="metric-label">💰 Gesamt Investiert</div>
                <div class="metric-value">{stats['total_invested']:,.2f} {self.currency}</div>
                <div class="metric-subtext">Alle Assets</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">📈 Realisierte Gewinne</div>
                <div class="metric-value {'positive' if stats['total_realized_gains'] >= 0 else 'negative'}">
                    {stats['total_realized_gains']:+,.2f} {self.currency}
                </div>
                <div class="metric-subtext">Brutto</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">💎 Unrealisierte Gewinne</div>
                <div class="metric-value {'positive' if stats['unrealized_gains'] >= 0 else 'negative'}">
                    {stats['unrealized_gains']:+,.2f} {self.currency}
                </div>
                <div class="metric-subtext">Aktuelle Positionen</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">🎯 Gesamt-Gewinne</div>
                <div class="metric-value {'positive' if stats['total_gains'] >= 0 else 'negative'}">
                    {stats['total_gains']:+,.2f} {self.currency}
                </div>
                <div class="metric-subtext">Realisiert + Unrealisiert</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">📊 Gesamt-Rendite</div>
                <div class="metric-value {'positive' if stats['total_return_pct'] >= 0 else 'negative'}">
                    {stats['total_return_pct']:+,.1f}%
                </div>
                <div class="metric-subtext">ROI</div>
            </div>

            <div class="metric-card">
                <div class="metric-label">💰 Aktueller Wert</div>
                <div class="metric-value">{stats['remaining_shares_value']:,.2f} {self.currency}</div>
                <div class="metric-subtext">Alle Positionen</div>
            </div>
        </div>

        <div class="section">
            <h2>📊 Asset-Übersicht</h2>
            <table>
                <thead>
                    <tr>
                        <th>Asset</th>
                        <th>Investiert</th>
                        <th>Entnommen</th>
                        <th>Realisiert</th>
                        <th>Unrealisiert</th>
                        <th>Gesamt</th>
                        <th>Rendite</th>
                    </tr>
                </thead>
                <tbody>
                    {assets_html}
                </tbody>
            </table>
        </div>
    </div>
</body>
</html>"""

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return output_file


def analyze_multi_asset_portfolio(csv_file_path, current_prices=None, currency="EUR"):
    """
    Analysiert Portfolio mit mehreren Assets

    Args:
        csv_file_path: CSV-Datei mit Company-Spalte oder Dict {company: csv_path}
        current_prices: Dictionary {company: price}
        currency: Währung (Standard: EUR)

    Returns:
        MultiAssetPortfolioAnalyzer: Analyzer mit allen Assets
    """
    analyzer = MultiAssetPortfolioAnalyzer(csv_file_path, current_prices, currency)
    analyzer.print_summary_report()
    return analyzer


# COMMAND-LINE INTERFACE
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description='Portfolio FIFO Analyzer - Automatische FIFO-Analyse für Aktien-Portfolios',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Beispiele:
  %(prog)s examples/palantir_example.csv 35.00 EUR
  %(prog)s csvs/meine_aktien.csv 150.50 USD --html output/report.html
  %(prog)s examples/nvidia_example.csv 485.00 EUR --html

Für weitere Informationen siehe README.md
        """
    )

    parser.add_argument('csv_file', help='Pfad zur CSV-Datei mit Transaktionen')
    parser.add_argument('current_price', type=float, nargs='?', default=None,
                       help='Aktueller Aktienkurs (optional)')
    parser.add_argument('currency', nargs='?', default='EUR',
                       help='Währung (Standard: EUR)')
    parser.add_argument('--html', nargs='?', const='output/portfolio_report.html',
                       metavar='OUTPUT_FILE',
                       help='Generiere HTML-Report (optional: Pfad angeben)')
    parser.add_argument('--excel', nargs='?', const='output/portfolio_report.xlsx',
                       metavar='OUTPUT_FILE',
                       help='Generiere Excel-Report (optional: Pfad angeben)')
    parser.add_argument('--pdf', nargs='?', const='output/portfolio_report.pdf',
                       metavar='OUTPUT_FILE',
                       help='Generiere PDF-Report (optional: Pfad angeben)')

    args = parser.parse_args()

    print("🚀 PORTFOLIO FIFO ANALYZER")
    print("=" * 50)

    try:
        # Analyse durchführen
        analyzer = analyze_portfolio_from_csv(
            args.csv_file,
            args.current_price,
            args.currency
        )

        # HTML-Report generieren falls gewünscht
        if args.html:
            report_file = analyzer.generate_html_report(args.html)
            print(f"\n📄 HTML-Report erstellt: {report_file}")

        # Excel-Report generieren falls gewünscht
        if args.excel:
            report_file = analyzer.generate_excel_report(args.excel)
            print(f"\n📊 Excel-Report erstellt: {report_file}")

        # PDF-Report generieren falls gewünscht
        if args.pdf:
            report_file = analyzer.generate_pdf_report(args.pdf)
            print(f"\n📑 PDF-Report erstellt: {report_file}")

        print(f"\n✅ Analyse abgeschlossen!")

    except FileNotFoundError:
        print(f"\n❌ CSV-Datei nicht gefunden: {args.csv_file}")
        print("💡 Bitte den korrekten Pfad zur CSV-Datei angeben")
        exit(1)
    except Exception as e:
        print(f"\n❌ Fehler bei der Analyse: {e}")
        print("💡 Bitte CSV-Format und Daten überprüfen")
        import traceback
        traceback.print_exc()
        exit(1)
